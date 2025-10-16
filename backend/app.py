from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
import os
import bcrypt
import json
import random
import string
from datetime import datetime, timedelta, timezone
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import io

from config import Config

app = Flask(__name__)
app.config.from_object(Config)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
# Normalize Gmail app password (allow spaces in .env but strip for SMTP auth)
if isinstance(app.config.get('MAIL_PASSWORD'), str):
    app.config['MAIL_PASSWORD'] = app.config['MAIL_PASSWORD'].replace(' ', '')

mail = Mail(app)
CORS(app)

# Time helpers
def utcnow_naive():
    """Return current UTC time as a timezone-naive datetime for DB consistency."""
    return datetime.now(timezone.utc).replace(tzinfo=None)

def sync_admin_credentials():
    """Sync hardcoded admin credentials with database on startup."""
    try:
        for employee_id, creds in ADMIN_CREDENTIALS.items():
            admin = Admin.query.filter_by(employee_id=employee_id).first()
            if not admin:
                # Create admin record if it doesn't exist
                admin = Admin(
                    name=f"Admin - {creds['branch']}",
                    employee_id=employee_id,
                    email=f"{employee_id}@srit.ac.in",
                    branch=creds['branch'],
                    password=hash_password(creds['password'])
                )
                db.session.add(admin)
                print(f"Created admin record for {employee_id}")
            else:
                # Update hardcoded password with database password if it exists
                # This ensures database password takes precedence
                if admin.password:
                    # Don't update if admin has a custom password in database
                    pass
                else:
                    # Set initial password from hardcoded credentials
                    admin.password = hash_password(creds['password'])
                    print(f"Set initial password for {employee_id}")
        
        db.session.commit()
        print("Admin credentials synced with database")
    except Exception as e:
        print(f"Error syncing admin credentials: {e}")

# Database Models
class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    rollnumber = db.Column(db.String(20), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(15))
    gender = db.Column(db.String(10), db.CheckConstraint("gender IN ('Male', 'Female', 'Other')"))
    branch = db.Column(db.String(100))
    section = db.Column(db.String(10))
    year = db.Column(db.String(10))
    password = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=utcnow_naive)

class Admin(db.Model):
    __tablename__ = 'admins'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(15))
    gender = db.Column(db.String(10), db.CheckConstraint("gender IN ('Male', 'Female', 'Other')"))
    branch = db.Column(db.String(100))
    password = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=utcnow_naive)

class AdminMessage(db.Model):
    __tablename__ = 'admin_messages'
    id = db.Column(db.Integer, primary_key=True)
    admin_id = db.Column(db.Integer, db.ForeignKey('admins.id'), nullable=False)
    subject = db.Column(db.String(255), nullable=False)
    body = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=utcnow_naive)

class Certificate(db.Model):
    __tablename__ = 'certificates'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    certificate_name = db.Column(db.String(200))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    name = db.Column(db.String(100))
    email = db.Column(db.String(120))
    branch = db.Column(db.String(100))
    year = db.Column(db.String(10))
    event_type = db.Column(db.String(100))
    file_path = db.Column(db.String(255))
    status = db.Column(db.String(20), db.CheckConstraint("status IN ('Pending', 'Approved', 'Rejected')"), default='Pending')
    uploaded_at = db.Column(db.DateTime, default=utcnow_naive)

class OTP(db.Model):
    __tablename__ = 'otps'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)
    otp_code = db.Column(db.String(6), nullable=False)
    purpose = db.Column(db.String(20), nullable=False)  # 'registration' or 'reset_password'
    created_at = db.Column(db.DateTime, default=utcnow_naive)
    expires_at = db.Column(db.DateTime, nullable=False)
    is_used = db.Column(db.Boolean, default=False)

class Form(db.Model):
    __tablename__ = 'forms'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    admin_id = db.Column(db.Integer, db.ForeignKey('admins.id'), nullable=False)
    branch = db.Column(db.String(100), nullable=False)
    deadline = db.Column(db.DateTime, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=utcnow_naive)
    form_fields = db.Column(db.Text)  # JSON string containing form fields

class FormResponse(db.Model):
    __tablename__ = 'form_responses'
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('forms.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    responses = db.Column(db.Text)  # JSON string containing responses
    submitted_at = db.Column(db.DateTime, default=utcnow_naive)

# Utility functions
def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def check_password(password, hashed):
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def verify_password(password, hashed):
    """Alias for check_password for clarity in login logic."""
    return check_password(password, hashed)

def validate_email(email):
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_password(password):
    # At least 8 characters, 1 uppercase, 1 lowercase, 1 number
    if len(password) < 8:
        return False, "Password must be at least 8 characters long"
    if not any(c.isupper() for c in password):
        return False, "Password must contain at least one uppercase letter"
    if not any(c.islower() for c in password):
        return False, "Password must contain at least one lowercase letter"
    if not any(c.isdigit() for c in password):
        return False, "Password must contain at least one number"
    return True, "Password is valid"

def validate_rollnumber(rollnumber):
    # 10 alphanumeric characters
    import re
    pattern = r'^[A-Za-z0-9]{10}$'
    return re.match(pattern, rollnumber) is not None

def generate_otp():
    """Generate a 6-digit OTP"""
    return ''.join(random.choices(string.digits, k=6))

def send_otp_email(email, otp_code, purpose):
    """Send OTP email to user"""
    try:
        if purpose == 'registration':
            subject = "SAT Portal - Email Verification OTP"
            body = f"""
            <html>
            <body>
                <h2>🎓 SAT Portal - Email Verification</h2>
                <p>Thank you for registering with SAT Portal!</p>
                <p>Your verification OTP is: <strong style="font-size: 24px; color: #007bff;">{otp_code}</strong></p>
                <p>This OTP will expire in 10 minutes.</p>
                <p>If you didn't request this registration, please ignore this email.</p>
                <br>
                <p>Best regards,<br>SAT Portal Team</p>
            </body>
            </html>
            """
        else:  # reset_password
            subject = "SAT Portal - Password Reset OTP"
            body = f"""
            <html>
            <body>
                <h2>🔐 SAT Portal - Password Reset</h2>
                <p>You requested a password reset for your SAT Portal account.</p>
                <p>Your reset OTP is: <strong style="font-size: 24px; color: #007bff;">{otp_code}</strong></p>
                <p>This OTP will expire in 10 minutes.</p>
                <p>If you didn't request this reset, please ignore this email.</p>
                <br>
                <p>Best regards,<br>SAT Portal Team</p>
            </body>
            </html>
            """
        
        sender_address = app.config.get('MAIL_DEFAULT_SENDER', app.config.get('MAIL_USERNAME'))
        msg = Message(subject, sender=sender_address, recipients=[email])
        msg.html = body
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Error sending email via Flask-Mail: {e}")
        # Fallback: try direct SMTP (first TLS 587, then SSL 465)
        try:
            username = app.config.get('MAIL_USERNAME')
            password = app.config.get('MAIL_PASSWORD')
            server_host = app.config.get('MAIL_SERVER', 'smtp.gmail.com')
            sender_address = app.config.get('MAIL_DEFAULT_SENDER', username)

            if not username or not password:
                print("SMTP fallback aborted: MAIL_USERNAME or MAIL_PASSWORD not configured")
                return False

            # Prepare MIME message
            mime_msg = MIMEText(body, 'html')
            mime_msg['Subject'] = subject
            mime_msg['From'] = formataddr(("SAT Portal", sender_address))
            mime_msg['To'] = email

            # Try TLS
            try:
                with smtplib.SMTP(server_host, 587, timeout=10) as smtp:
                    smtp.ehlo()
                    smtp.starttls()
                    smtp.login(username, password)
                    smtp.sendmail(sender_address, [email], mime_msg.as_string())
                    return True
            except Exception as e_tls:
                print(f"TLS send failed, trying SSL: {e_tls}")

            # Try SSL
            try:
                with smtplib.SMTP_SSL(server_host, 465, timeout=10) as smtp:
                    smtp.login(username, password)
                    smtp.sendmail(sender_address, [email], mime_msg.as_string())
                    return True
            except Exception as e_ssl:
                print(f"SSL send failed: {e_ssl}")
                return False
        except Exception as e_fb:
            print(f"SMTP fallback error: {e_fb}")
            return False

def create_otp_record(email, purpose):
    """Create OTP record in database"""
    # Delete any existing unused OTPs for this email and purpose
    OTP.query.filter_by(email=email, purpose=purpose, is_used=False).delete()
    
    otp_code = generate_otp()
    # Store expiration as UTC-naive to match DB timestamps
    expires_at = utcnow_naive() + timedelta(minutes=10)
    
    otp_record = OTP(
        email=email,
        otp_code=otp_code,
        purpose=purpose,
        expires_at=expires_at
    )
    
    db.session.add(otp_record)
    db.session.commit()
    
    return otp_code

def verify_otp(email, otp_code, purpose):
    """Verify OTP from database"""
    otp_record = OTP.query.filter_by(
        email=email, 
        otp_code=otp_code, 
        purpose=purpose, 
        is_used=False
    ).first()
    
    if not otp_record:
        return False, "Invalid OTP"
    
    # Compare using UTC-naive current time to match stored naive timestamp
    current_utc_naive = utcnow_naive()
    if current_utc_naive > otp_record.expires_at:
        return False, "OTP has expired"
    
    # Mark OTP as used
    otp_record.is_used = True
    db.session.commit()
    
    return True, "OTP verified successfully"

# Admin credentials
ADMIN_CREDENTIALS = {
    'admin@cse': {'password': 'Cse@srit', 'branch': 'COMPUTER SCIENCE AND ENGINEERING'},
    'admin@csd': {'password': 'Csd@srit', 'branch': 'COMPUTER SCIENCE AND ENGINEERING [DATA SCIENCE]'},
    'admin@csm': {'password': 'Csm@srit', 'branch': 'COMPUTER SCIENCE AND ENGINEERING [AIML]'},
    'admin@ece': {'password': 'Ece@srit', 'branch': 'ELECTRICAL AND ELECTRONICS OF COMMUNICATION ENGINEERING'},
    'admin@eee': {'password': 'Eee@srit', 'branch': 'ELECTRICAL AND ELECTRONICS ENGINEERING'},
    'admin@civ': {'password': 'Civ@srit', 'branch': 'CIVIL ENGINEERING'},
    'admin@mech': {'password': 'Mech@srit', 'branch': 'MECHANICAL ENGINEERING'},
    # Super admin credential
    'super@admin': {'password': 'Superadmin@srit', 'branch': 'SUPER ADMIN', 'role': 'superadmin'}
}

# Routes
@app.route('/api/auth/student/send-otp', methods=['POST'])
def send_registration_otp():
    try:
        data = request.get_json()
        
        if not data.get('email'):
            return jsonify({'error': 'Email is required'}), 400
        
        # Email validation
        if not validate_email(data['email']):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Check if email already registered
        if Student.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already registered'}), 400
        
        # Generate and send OTP
        otp_code = create_otp_record(data['email'], 'registration')
        
        if send_otp_email(data['email'], otp_code, 'registration'):
            return jsonify({'message': 'OTP sent successfully to your email'}), 200
        else:
            # Development fallback: allow proceeding if email fails in debug
            if app.debug:
                print(f"DEV MODE: Registration OTP for {data['email']} is {otp_code}")
                return jsonify({'message': 'OTP send failed, but development fallback applied. Check server logs for OTP.'}), 200
            return jsonify({'error': 'Failed to send OTP. Please try again.'}), 500
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/student/register', methods=['POST'])
def student_register():
    try:
        data = request.get_json()
        
        # Validation
        if not all(key in data for key in ['name', 'rollnumber', 'email', 'password', 'phone', 'gender', 'branch', 'section', 'year', 'otp']):
            return jsonify({'error': 'All fields including OTP are required'}), 400
        
        # Email validation
        if not validate_email(data['email']):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Roll number validation
        if not validate_rollnumber(data['rollnumber']):
            return jsonify({'error': 'Roll number must be exactly 10 alphanumeric characters'}), 400
        
        # Password validation
        is_valid, message = validate_password(data['password'])
        if not is_valid:
            return jsonify({'error': message}), 400
        
        # Check if student already exists
        if Student.query.filter_by(rollnumber=data['rollnumber']).first():
            return jsonify({'error': 'Roll number already registered'}), 400
        
        if Student.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already registered'}), 400
        
        # Verify OTP
        is_valid_otp, otp_message = verify_otp(data['email'], data['otp'], 'registration')
        if not is_valid_otp:
            return jsonify({'error': otp_message}), 400
        
        # Create new student
        hashed_password = hash_password(data['password'])
        new_student = Student(
            name=data['name'],
            rollnumber=data['rollnumber'],
            email=data['email'],
            phone=data['phone'],
            gender=data['gender'],
            branch=data['branch'],
            section=data['section'],
            year=data['year'],
            password=hashed_password
        )
        
        db.session.add(new_student)
        db.session.commit()
        
        return jsonify({'message': 'Student registered successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/student/login', methods=['POST'])
def student_login():
    try:
        data = request.get_json()
        
        if not all(key in data for key in ['rollnumber', 'password']):
            return jsonify({'error': 'Roll number and password are required'}), 400
        
        student = Student.query.filter_by(rollnumber=data['rollnumber']).first()
        
        if not student or not check_password(data['password'], student.password):
            return jsonify({'error': 'Invalid roll number or password'}), 401
        
        return jsonify({
            'message': 'Login successful',
            'student': {
                'id': student.id,
                'name': student.name,
                'rollnumber': student.rollnumber,
                'email': student.email,
                'phone': student.phone,
                'gender': student.gender,
                'branch': student.branch,
                'section': student.section,
                'year': student.year
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/admin/login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json()
        
        if not all(key in data for key in ['employee_id', 'password']):
            return jsonify({'error': 'Employee ID and password are required'}), 400
        
        # First check if admin exists in database (for updated passwords)
        admin = Admin.query.filter_by(employee_id=data['employee_id']).first()
        
        if admin:
            # Admin exists in database, check stored password
            if verify_password(data['password'], admin.password):
                # Get role from ADMIN_CREDENTIALS for super admin check
                admin_cred = ADMIN_CREDENTIALS.get(data['employee_id'], {})
                return jsonify({
                    'message': 'Login successful',
                    'admin': {
                        'id': admin.id,
                        'name': admin.name,
                        'employee_id': admin.employee_id,
                        'email': admin.email,
                        'phone': admin.phone,
                        'gender': admin.gender,
                        'branch': admin.branch,
                        'role': admin_cred.get('role', 'admin')
                    }
                }), 200
            else:
                return jsonify({'error': 'Invalid password'}), 401
        
        # If not in database, check hardcoded credentials (for initial setup)
        elif data['employee_id'] in ADMIN_CREDENTIALS:
            admin_cred = ADMIN_CREDENTIALS[data['employee_id']]
            if data['password'] == admin_cred['password']:
                # Create admin record in database
                admin = Admin(
                    name=f"Admin - {admin_cred['branch']}",
                    employee_id=data['employee_id'],
                    email=f"{data['employee_id']}@srit.ac.in",
                    branch=admin_cred['branch'],
                    password=hash_password(data['password'])
                )
                db.session.add(admin)
                db.session.commit()
                
                return jsonify({
                    'message': 'Login successful',
                    'admin': {
                        'id': admin.id,
                        'name': admin.name,
                        'employee_id': admin.employee_id,
                        'email': admin.email,
                        'phone': admin.phone,
                        'gender': admin.gender,
                        'branch': admin.branch,
                        'role': admin_cred.get('role', 'admin')
                    }
                }), 200
            else:
                return jsonify({'error': 'Invalid password'}), 401
        else:
            return jsonify({'error': 'Invalid employee ID'}), 401
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/student/forgot-password', methods=['POST'])
def forgot_password():
    try:
        data = request.get_json()
        
        if not data.get('email'):
            return jsonify({'error': 'Email is required'}), 400
        
        # Email validation
        if not validate_email(data['email']):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Check if email exists
        student = Student.query.filter_by(email=data['email']).first()
        if not student:
            return jsonify({'error': 'No account found with this email address'}), 404
        
        # Generate and send OTP
        otp_code = create_otp_record(data['email'], 'reset_password')
        
        if send_otp_email(data['email'], otp_code, 'reset_password'):
            return jsonify({'message': 'Password reset OTP sent successfully to your email'}), 200
        else:
            # Development fallback: allow proceeding if email fails in debug
            if app.debug:
                print(f"DEV MODE: Password reset OTP for {data['email']} is {otp_code}")
                return jsonify({'message': 'OTP send failed, but development fallback applied. Check server logs for OTP.'}), 200
            return jsonify({'error': 'Failed to send OTP. Please try again.'}), 500
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/student/reset-password', methods=['POST'])
def reset_password():
    try:
        data = request.get_json()
        
        if not all(key in data for key in ['email', 'otp', 'new_password']):
            return jsonify({'error': 'Email, OTP, and new password are required'}), 400
        
        # Email validation
        if not validate_email(data['email']):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Password validation
        is_valid, message = validate_password(data['new_password'])
        if not is_valid:
            return jsonify({'error': message}), 400
        
        # Check if email exists
        student = Student.query.filter_by(email=data['email']).first()
        if not student:
            return jsonify({'error': 'No account found with this email address'}), 404
        
        # Verify OTP
        is_valid_otp, otp_message = verify_otp(data['email'], data['otp'], 'reset_password')
        if not is_valid_otp:
            return jsonify({'error': otp_message}), 400
        
        # Update password
        student.password = hash_password(data['new_password'])
        db.session.commit()
        
        return jsonify({'message': 'Password reset successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/profile', methods=['GET'])
def get_student_profile():
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student ID is required'}), 400
        
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        return jsonify({
            'id': student.id,
            'name': student.name,
            'rollnumber': student.rollnumber,
            'email': student.email,
            'phone': student.phone,
            'gender': student.gender,
            'branch': student.branch,
            'section': student.section,
            'year': student.year
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/profile', methods=['PUT'])
def update_student_profile():
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'error': 'Student ID is required'}), 400
        
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Update fields
        if 'name' in data:
            student.name = data['name']
        if 'email' in data:
            if not validate_email(data['email']):
                return jsonify({'error': 'Invalid email format'}), 400
            student.email = data['email']
        if 'phone' in data:
            student.phone = data['phone']
        if 'gender' in data:
            student.gender = data['gender']
        if 'branch' in data:
            student.branch = data['branch']
        if 'section' in data:
            student.section = data['section']
        if 'year' in data:
            student.year = data['year']
        
        db.session.commit()
        
        return jsonify({'message': 'Profile updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/certificate/upload', methods=['POST'])
def upload_certificate():
    try:
        student_id = request.form.get('student_id')
        event_type = request.form.get('event_type')
        certificate_name = request.form.get('certificate_name', '')
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')
        
        if not all([student_id, event_type]):
            return jsonify({'error': 'Student ID and event type are required'}), 400
        
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        if 'certificate' not in request.files:
            return jsonify({'error': 'Certificate file is required'}), 400
        
        file = request.files['certificate']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Save file
        filename = f"{student.rollnumber}_{event_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Parse dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Create certificate record
        certificate = Certificate(
            student_id=student.id,
            certificate_name=certificate_name,
            start_date=start_date_obj,
            end_date=end_date_obj,
            name=student.name,
            email=student.email,
            branch=student.branch,
            year=student.year,
            event_type=event_type,
            file_path=file_path,
            status='Pending'
        )
        
        db.session.add(certificate)
        db.session.commit()
        
        return jsonify({'message': 'Certificate uploaded successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/certificates', methods=['GET'])
def get_student_certificates():
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student ID is required'}), 400
        
        certificates = Certificate.query.filter_by(student_id=student_id).all()
        
        certificate_list = []
        for cert in certificates:
            certificate_list.append({
                'id': cert.id,
                'certificate_name': cert.certificate_name,
                'event_type': cert.event_type,
                'start_date': cert.start_date.strftime('%Y-%m-%d') if cert.start_date else None,
                'end_date': cert.end_date.strftime('%Y-%m-%d') if cert.end_date else None,
                'status': cert.status,
                'uploaded_at': cert.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
                'file_path': cert.file_path
            })
        
        return jsonify({'certificates': certificate_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/certificate/<int:certificate_id>/download', methods=['GET'])
def download_certificate(certificate_id):
    try:
        certificate = db.session.get(Certificate, certificate_id)
        if not certificate:
            return jsonify({'error': 'Certificate not found'}), 404
        
        if not os.path.exists(certificate.file_path):
            return jsonify({'error': 'Certificate file not found'}), 404
        
        return send_file(
            certificate.file_path,
            as_attachment=True,
            download_name=os.path.basename(certificate.file_path),
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/certificate/<int:certificate_id>/view', methods=['GET'])
def view_certificate(certificate_id):
    try:
        certificate = db.session.get(Certificate, certificate_id)
        if not certificate:
            return jsonify({'error': 'Certificate not found'}), 404
        
        if not os.path.exists(certificate.file_path):
            return jsonify({'error': 'Certificate file not found'}), 404
        
        return send_file(
            certificate.file_path,
            as_attachment=False,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/dashboard', methods=['GET'])
def admin_dashboard():
    try:
        # Get admin branch from request headers or query params
        admin_branch = request.args.get('branch') or request.headers.get('X-Admin-Branch')
        
        # Build query with branch filter if provided
        query = Certificate.query
        if admin_branch:
            query = query.filter_by(branch=admin_branch)
        
        # Get statistics filtered by branch
        total_certificates = query.count()
        pending_certificates = query.filter_by(status='Pending').count()
        approved_certificates = query.filter_by(status='Approved').count()
        rejected_certificates = query.filter_by(status='Rejected').count()
        
        # Get recent certificates filtered by branch
        recent_certificates = query.order_by(Certificate.uploaded_at.desc()).limit(10).all()
        
        recent_list = []
        for cert in recent_certificates:
            student = db.session.get(Student, cert.student_id)
            recent_list.append({
                'id': cert.id,
                'student_name': cert.name,
                'rollnumber': student.rollnumber if student else 'N/A',
                'event_type': cert.event_type,
                'status': cert.status,
                'uploaded_at': cert.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'statistics': {
                'total_certificates': total_certificates,
                'pending_certificates': pending_certificates,
                'approved_certificates': approved_certificates,
                'rejected_certificates': rejected_certificates
            },
            'recent_certificates': recent_list
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/certificates', methods=['GET'])
def get_admin_certificates():
    try:
        # Get filter parameters
        year = request.args.get('year')
        branch = request.args.get('branch')
        section = request.args.get('section')
        event_type = request.args.get('event_type')
        status = request.args.get('status')
        
        query = Certificate.query
        
        if year:
            query = query.filter_by(year=year)
        if branch:
            query = query.filter_by(branch=branch)
        if event_type:
            query = query.filter_by(event_type=event_type)
        if status:
            query = query.filter_by(status=status)
        
        certificates = query.order_by(Certificate.uploaded_at.desc()).all()
        
        certificate_list = []
        for cert in certificates:
            student = db.session.get(Student, cert.student_id)
            certificate_list.append({
                'id': cert.id,
                'student_name': cert.name,
                'rollnumber': student.rollnumber if student else 'N/A',
                'email': cert.email,
                'branch': cert.branch,
                'section': student.section if student else 'N/A',
                'year': cert.year,
                'certificate_name': cert.certificate_name,
                'event_type': cert.event_type,
                'start_date': cert.start_date.strftime('%Y-%m-%d') if cert.start_date else None,
                'end_date': cert.end_date.strftime('%Y-%m-%d') if cert.end_date else None,
                'status': cert.status,
                'uploaded_at': cert.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
                'file_path': cert.file_path
            })
        
        return jsonify({'certificates': certificate_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/students', methods=['GET'])
def get_admin_students():
    try:
        branch = request.args.get('branch')
        year = request.args.get('year')
        section = request.args.get('section')
        
        query = Student.query
        if branch:
            query = query.filter_by(branch=branch)
        if year:
            query = query.filter_by(year=year)
        if section:
            query = query.filter_by(section=section)
        
        students = query.all()
        
        student_list = []
        for student in students:
            student_list.append({
                'id': student.id,
                'name': student.name,
                'rollnumber': student.rollnumber,
                'email': student.email,
                'phone': student.phone,
                'gender': student.gender,
                'branch': student.branch,
                'section': student.section,
                'year': student.year
            })
        
        return jsonify({'students': student_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/students/download', methods=['GET'])
def download_students_excel():
    try:
        branch = request.args.get('branch')
        year = request.args.get('year')
        section = request.args.get('section')
        
        query = Student.query
        if branch:
            query = query.filter_by(branch=branch)
        if year:
            query = query.filter_by(year=year)
        if section:
            query = query.filter_by(section=section)
        
        students = query.all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students Report"
        
        # Headers
        headers = ['ID', 'Name', 'Roll Number', 'Email', 'Phone', 'Gender', 'Branch', 'Section', 'Year', 'Created At']
        ws.append(headers)
        
        # Data
        for student in students:
            ws.append([
                student.id,
                student.name,
                student.rollnumber,
                student.email,
                student.phone or 'N/A',
                student.gender or 'N/A',
                student.branch or 'N/A',
                student.section or 'N/A',
                student.year or 'N/A',
                student.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='students_report.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/certificate/status', methods=['PUT'])
def update_certificate_status():
    try:
        data = request.get_json()
        certificate_id = data.get('certificate_id')
        status = data.get('status')
        
        if not all([certificate_id, status]):
            return jsonify({'error': 'Certificate ID and status are required'}), 400
        
        if status not in ['Pending', 'Approved', 'Rejected']:
            return jsonify({'error': 'Invalid status'}), 400
        
        certificate = db.session.get(Certificate, certificate_id)
        if not certificate:
            return jsonify({'error': 'Certificate not found'}), 404
        
        certificate.status = status
        db.session.commit()
        
        return jsonify({'message': 'Certificate status updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/certificate/bulk-status', methods=['PUT'])
def bulk_update_certificate_status():
    try:
        data = request.get_json()
        certificate_ids = data.get('certificate_ids', [])
        status = data.get('status')
        
        if not certificate_ids or not status:
            return jsonify({'error': 'Certificate IDs and status are required'}), 400
        
        if status not in ['Pending', 'Approved', 'Rejected']:
            return jsonify({'error': 'Invalid status'}), 400
        
        certificates = Certificate.query.filter(Certificate.id.in_(certificate_ids)).all()
        for certificate in certificates:
            certificate.status = status
        
        db.session.commit()
        
        return jsonify({'message': f'Updated {len(certificates)} certificates successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/analytics', methods=['GET'])
def get_analytics():
    try:
        # Get filter parameters
        year = request.args.get('year')
        branch = request.args.get('branch')
        event_type = request.args.get('event_type')
        
        # Build base query with branch filter if provided
        base_query = Certificate.query
        if branch:
            base_query = base_query.filter_by(branch=branch)
        
        # Apply additional filters
        if year:
            base_query = base_query.filter_by(year=year)
        if event_type:
            base_query = base_query.filter_by(event_type=event_type)
        
        # Branch-wise statistics (only if no specific branch filter)
        if not branch:
            branch_stats = db.session.query(
                Certificate.branch,
                db.func.count(Certificate.id).label('count')
            ).group_by(Certificate.branch).all()
        else:
            branch_stats = [{'branch': branch, 'count': base_query.count()}]
        
        # Year-wise statistics
        year_stats = db.session.query(
            Certificate.year,
            db.func.count(Certificate.id).label('count')
        )
        if branch:
            year_stats = year_stats.filter(Certificate.branch == branch)
        if year:
            year_stats = year_stats.filter(Certificate.year == year)
        if event_type:
            year_stats = year_stats.filter(Certificate.event_type == event_type)
        year_stats = year_stats.group_by(Certificate.year).all()
        
        # Event type statistics
        event_stats = db.session.query(
            Certificate.event_type,
            db.func.count(Certificate.id).label('count')
        )
        if branch:
            event_stats = event_stats.filter(Certificate.branch == branch)
        if year:
            event_stats = event_stats.filter(Certificate.year == year)
        if event_type:
            event_stats = event_stats.filter(Certificate.event_type == event_type)
        event_stats = event_stats.group_by(Certificate.event_type).all()
        
        # Status statistics
        status_stats = db.session.query(
            Certificate.status,
            db.func.count(Certificate.id).label('count')
        )
        if branch:
            status_stats = status_stats.filter(Certificate.branch == branch)
        if year:
            status_stats = status_stats.filter(Certificate.year == year)
        if event_type:
            status_stats = status_stats.filter(Certificate.event_type == event_type)
        status_stats = status_stats.group_by(Certificate.status).all()
        
        return jsonify({
            'branch_stats': [{'branch': stat.branch, 'count': stat.count} for stat in branch_stats] if not branch else [{'branch': branch, 'count': base_query.count()}],
            'year_stats': [{'year': stat.year, 'count': stat.count} for stat in year_stats],
            'event_stats': [{'event_type': stat.event_type, 'count': stat.count} for stat in event_stats],
            'status_stats': [{'status': stat.status, 'count': stat.count} for stat in status_stats]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/report', methods=['GET'])
def generate_report():
    try:
        report_type = request.args.get('type', 'excel')  # excel or pdf
        year = request.args.get('year')
        branch = request.args.get('branch')
        status = request.args.get('status')
        
        # Check if at least one filter is applied
        if not any([year, branch, status]):
            return jsonify({'error': 'Please select at least one filter (Year, Branch, or Status) before downloading the report'}), 400
        
        query = Certificate.query
        
        if year:
            query = query.filter_by(year=year)
        if branch:
            query = query.filter_by(branch=branch)
        if status:
            query = query.filter_by(status=status)
        
        certificates = query.all()
        
        if report_type == 'excel':
            return generate_excel_report(certificates)
        else:
            return generate_pdf_report(certificates)
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_excel_report(certificates):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificates Report"
    
    # Headers
    headers = ['ID', 'Student Name', 'Roll Number', 'Email', 'Branch', 'Year', 'Certificate Name', 'Event Type', 'Start Date', 'End Date', 'Status', 'Uploaded At', 'Certificate PDF']
    ws.append(headers)
    
    # Data
    for cert in certificates:
        student = db.session.get(Student, cert.student_id)
        # Create a proper download link that will work when the Excel file is opened
        certificate_pdf_link = f"=HYPERLINK(\"http://localhost:5000/api/student/certificate/{cert.id}/download\", \"Download Certificate\")" if cert.file_path else "N/A"
        ws.append([
            cert.id,
            cert.name,
            student.rollnumber if student else 'N/A',
            cert.email,
            cert.branch,
            cert.year,
            cert.certificate_name or cert.event_type,
            cert.event_type,
            cert.start_date.strftime('%Y-%m-%d') if cert.start_date else 'N/A',
            cert.end_date.strftime('%Y-%m-%d') if cert.end_date else 'N/A',
            cert.status,
            cert.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
            certificate_pdf_link
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='certificates_report.xlsx'
    )

def generate_pdf_report(certificates):
    buffer = io.BytesIO()
    
    # Create PDF with watermark in landscape mode
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Title with smaller font
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontSize = 14
    title = Paragraph("Certificate Reports", title_style)
    elements.append(title)
    
    # Data with reduced columns for better fit
    data = [['ID', 'Student Name', 'Roll Number', 'Email', 'Branch', 'Year', 'Event Type', 'Status', 'Uploaded At']]
    
    for cert in certificates:
        student = db.session.get(Student, cert.student_id)
        data.append([
            str(cert.id),
            cert.name[:20] + '...' if len(cert.name) > 20 else cert.name,  # Truncate long names
            student.rollnumber if student else 'N/A',
            cert.email[:25] + '...' if len(cert.email) > 25 else cert.email,  # Truncate long emails
            cert.branch[:15] + '...' if len(cert.branch) > 15 else cert.branch,  # Truncate long branch names
            str(cert.year),
            cert.event_type[:12] + '...' if len(cert.event_type) > 12 else cert.event_type,  # Truncate event type
            cert.status,
            cert.uploaded_at.strftime('%Y-%m-%d')  # Only date, not time
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller header font
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Much smaller data font
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Thinner grid
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF with watermark
    def add_watermark(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 60)
        canvas.setFillColor(colors.orange)
        canvas.setFillAlpha(0.3)
        canvas.rotate(45)
        canvas.drawString(200, 0, "SRIT")
        canvas.restoreState()
    
    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name='certificates_report.pdf'
    )

# Form Management API Endpoints

@app.route('/api/admin/forms', methods=['POST'])
def create_form():
    try:
        data = request.get_json()
        
        # Get admin info from session or token
        admin_id = data.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        # Create new form
        new_form = Form(
            title=data['title'],
            description=data.get('description', ''),
            admin_id=admin_id,
            branch=admin.branch,
            deadline=datetime.fromisoformat(data['deadline'].replace('Z', '+00:00')),
            form_fields=json.dumps(data['form_fields'])
        )
        
        db.session.add(new_form)
        db.session.commit()
        
        # Send email notifications to students in the same branch
        students = Student.query.filter_by(branch=admin.branch).all()
        
        for student in students:
            try:
                msg = Message(
                    subject=f'New Form Available: {new_form.title}',
                    sender=app.config['MAIL_USERNAME'],
                    recipients=[student.email]
                )
                msg.body = f"""
Dear {student.name},

Your department admin has created a new form: "{new_form.title}"

Description: {new_form.description}

Deadline: {new_form.deadline.strftime('%Y-%m-%d %H:%M:%S')}

Please log in to your student dashboard to fill out this form.

Best regards,
SRIT Admin Team
                """
                mail.send(msg)
            except Exception as e:
                print(f"Failed to send email to {student.email}: {str(e)}")
        
        return jsonify({
            'message': 'Form created successfully',
            'form_id': new_form.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms', methods=['GET'])
def get_admin_forms():
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        forms = Form.query.filter_by(admin_id=admin_id).order_by(Form.created_at.desc()).all()
        
        forms_data = []
        for form in forms:
            # Get response count
            response_count = FormResponse.query.filter_by(form_id=form.id).count()
            
            forms_data.append({
                'id': form.id,
                'title': form.title,
                'description': form.description,
                'branch': form.branch,
                'deadline': form.deadline.isoformat(),
                'is_active': form.is_active,
                'created_at': form.created_at.isoformat(),
                'response_count': response_count
            })
        
        return jsonify({'forms': forms_data}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/<int:form_id>/responses', methods=['GET'])
def get_form_responses(form_id):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        # Verify admin owns this form
        form = Form.query.filter_by(id=form_id, admin_id=admin_id).first()
        if not form:
            return jsonify({'error': 'Form not found or access denied'}), 404
        
        responses = FormResponse.query.filter_by(form_id=form_id).all()
        
        responses_data = []
        for response in responses:
            student = db.session.get(Student, response.student_id)
            responses_data.append({
                'id': response.id,
                'student_name': student.name if student else 'Unknown',
                'student_rollnumber': student.rollnumber if student else 'Unknown',
                'student_email': student.email if student else 'Unknown',
                'responses': json.loads(response.responses),
                'submitted_at': response.submitted_at.isoformat()
            })
        
        return jsonify({
            'form': {
                'id': form.id,
                'title': form.title,
                'description': form.description,
                'form_fields': json.loads(form.form_fields)
            },
            'responses': responses_data
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/<int:form_id>/responses/download', methods=['GET'])
def download_form_responses_excel(form_id):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        # Verify admin owns this form
        form = Form.query.filter_by(id=form_id, admin_id=admin_id).first()
        if not form:
            return jsonify({'error': 'Form not found or access denied'}), 404
        
        responses = FormResponse.query.filter_by(form_id=form_id).all()
        
        if not responses:
            return jsonify({'error': 'No responses found for this form'}), 404
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Form Responses"
        
        # Get form fields
        form_fields = json.loads(form.form_fields)
        
        # Create headers
        headers = ['Student Name', 'Roll Number', 'Student Email', 'Submission Date']
        for field in form_fields:
            headers.append(field['label'])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, response in enumerate(responses, 2):
            student = db.session.get(Student, response.student_id)
            responses_data = json.loads(response.responses)
            
            # Basic student info
            ws.cell(row=row, column=1, value=student.name if student else 'Unknown')
            ws.cell(row=row, column=2, value=student.rollnumber if student else 'Unknown')
            ws.cell(row=row, column=3, value=student.email if student else 'Unknown')
            ws.cell(row=row, column=4, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Form field responses
            for col, field in enumerate(form_fields, 5):
                field_id = str(field['id'])
                value = responses_data.get(field_id, '')
                
                # Handle different field types
                if field['type'] in ['checkbox', 'radio', 'select'] and isinstance(value, list):
                    value = ', '.join(value)
                elif field['type'] == 'file' and value:
                    # Create downloadable link for file uploads only if files exist
                    if isinstance(value, list):
                        # Multiple files
                        file_links = []
                        for i, filename in enumerate(value):
                            # Check if file exists
                            file_path = os.path.join('uploads', 'forms', str(form_id), str(response.id), filename)
                            if os.path.exists(file_path):
                                # Remove temp_ prefix for display
                                display_name = filename.replace('temp_', '') if filename.startswith('temp_') else filename
                                download_url = f"http://localhost:5000/api/admin/forms/{form_id}/responses/{response.id}/files/{filename}?admin_id={admin_id}"
                                file_links.append(f'=HYPERLINK("{download_url}","{display_name}")')
                            else:
                                # File doesn't exist, show filename without link
                                display_name = filename.replace('temp_', '') if filename.startswith('temp_') else filename
                                file_links.append(f'File not found: {display_name}')
                        value = ' | '.join(file_links)
                    else:
                        # Single file
                        # Check if file exists
                        file_path = os.path.join('uploads', 'forms', str(form_id), str(response.id), value)
                        if os.path.exists(file_path):
                            # Remove temp_ prefix for display
                            display_name = value.replace('temp_', '') if value.startswith('temp_') else value
                            download_url = f"http://localhost:5000/api/admin/forms/{form_id}/responses/{response.id}/files/{value}?admin_id={admin_id}"
                            value = f'=HYPERLINK("{download_url}","{display_name}")'
                        else:
                            # File doesn't exist, show filename without link
                            display_name = value.replace('temp_', '') if value.startswith('temp_') else value
                            value = f'File not found: {display_name}'
                
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'form_responses_{form.title.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/forms/<int:form_id>/response', methods=['GET'])
def get_student_form_response(form_id):
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student ID required'}), 400
        
        # Check if student has responded to this form
        response = FormResponse.query.filter_by(
            form_id=form_id,
            student_id=student_id
        ).first()
        
        if not response:
            return jsonify({'error': 'No response found for this form'}), 404
        
        return jsonify({
            'responses': json.loads(response.responses),
            'submitted_at': response.submitted_at.isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/<int:form_id>/responses/<int:response_id>/files/<path:filename>', methods=['GET'])
def download_form_file(form_id, response_id, filename):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        # Verify admin owns this form
        form = Form.query.filter_by(id=form_id, admin_id=admin_id).first()
        if not form:
            return jsonify({'error': 'Form not found or access denied'}), 404
        
        # Verify response exists for this form
        response = FormResponse.query.filter_by(id=response_id, form_id=form_id).first()
        if not response:
            return jsonify({'error': 'Response not found'}), 404
        
        # Construct file path
        file_path = os.path.join('uploads', 'forms', str(form_id), str(response_id), filename)
        
        # Check if file exists
        if not os.path.exists(file_path):
            # Check if directory exists
            response_dir = os.path.join('uploads', 'forms', str(form_id), str(response_id))
            if not os.path.exists(response_dir):
                return jsonify({'error': f'Response directory not found: {response_dir}'}), 404
            
            # List available files in the directory
            available_files = os.listdir(response_dir) if os.path.exists(response_dir) else []
            return jsonify({
                'error': f'File not found: {filename}',
                'available_files': available_files,
                'searched_path': file_path
            }), 404
        
        return send_file(file_path, as_attachment=True)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/forms', methods=['GET'])
def get_student_forms():
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student ID required'}), 400
        
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Get active forms for student's branch
        forms = Form.query.filter_by(
            branch=student.branch,
            is_active=True
        ).filter(Form.deadline > datetime.utcnow()).order_by(Form.created_at.desc()).all()
        
        forms_data = []
        for form in forms:
            # Check if student has already responded
            existing_response = FormResponse.query.filter_by(
                form_id=form.id,
                student_id=student_id
            ).first()
            
            forms_data.append({
                'id': form.id,
                'title': form.title,
                'description': form.description,
                'deadline': form.deadline.isoformat(),
                'created_at': form.created_at.isoformat(),
                'form_fields': json.loads(form.form_fields),
                'has_responded': existing_response is not None
            })
        
        return jsonify({'forms': forms_data}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/forms/<int:form_id>/submit', methods=['POST'])
def submit_form_response(form_id):
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        responses = data.get('responses')
        
        if not all([form_id, student_id, responses]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Check if form exists and is active
        form = Form.query.filter_by(id=form_id, is_active=True).first()
        if not form:
            return jsonify({'error': 'Form not found or inactive'}), 404
        
        # Check if deadline has passed
        if form.deadline < datetime.utcnow():
            return jsonify({'error': 'Form deadline has passed'}), 400
        
        # Check if student has already responded
        existing_response = FormResponse.query.filter_by(
            form_id=form_id,
            student_id=student_id
        ).first()
        
        if existing_response:
            return jsonify({'error': 'You have already submitted a response to this form'}), 400
        
        # Create new response
        new_response = FormResponse(
            form_id=form_id,
            student_id=student_id,
            responses=json.dumps(responses)
        )
        
        db.session.add(new_response)
        db.session.commit()
        
        # Move uploaded files from temp to final location
        temp_dir = os.path.join('uploads', 'forms', str(form_id), 'temp')
        final_dir = os.path.join('uploads', 'forms', str(form_id), str(new_response.id))
        
        if os.path.exists(temp_dir):
            os.makedirs(final_dir, exist_ok=True)
            for filename in os.listdir(temp_dir):
                if filename.startswith('temp_'):
                    # Move file to final location
                    src_path = os.path.join(temp_dir, filename)
                    dst_path = os.path.join(final_dir, filename)
                    os.rename(src_path, dst_path)
            
            # Remove temp directory
            try:
                os.rmdir(temp_dir)
            except OSError:
                pass  # Directory not empty, leave it
        
        return jsonify({'message': 'Form submitted successfully'}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/forms/<int:form_id>/upload-file', methods=['POST'])
def upload_form_file(form_id):
    try:
        student_id = request.form.get('student_id')
        field_id = request.form.get('field_id')
        
        if not all([form_id, student_id, field_id]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Check if form exists and is active
        form = Form.query.filter_by(id=form_id, is_active=True).first()
        if not form:
            return jsonify({'error': 'Form not found or inactive'}), 404
        
        # Check if deadline has passed
        if form.deadline < datetime.utcnow():
            return jsonify({'error': 'Form deadline has passed'}), 400
        
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file type (only PDF for now)
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Create upload directory structure
        upload_dir = os.path.join('uploads', 'forms', str(form_id), 'temp')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file with unique name
        filename = f"temp_{field_id}_{int(datetime.utcnow().timestamp())}_{file.filename}"
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        return jsonify({
            'message': 'File uploaded successfully',
            'filename': filename,
            'file_path': file_path
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/forms/notifications', methods=['GET'])
def get_form_notifications():
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student ID required'}), 400
        
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Get unresponded active forms
        forms = Form.query.filter_by(
            branch=student.branch,
            is_active=True
        ).filter(Form.deadline > datetime.utcnow()).all()
        
        unresponded_forms = []
        for form in forms:
            existing_response = FormResponse.query.filter_by(
                form_id=form.id,
                student_id=student_id
            ).first()
            
            if not existing_response:
                unresponded_forms.append({
                    'id': form.id,
                    'title': form.title,
                    'deadline': form.deadline.isoformat()
                })
        
        return jsonify({
            'unresponded_count': len(unresponded_forms),
            'forms': unresponded_forms
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/send-deadline-reminders', methods=['POST'])
def send_deadline_reminders():
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        # Get forms created by this admin that have deadlines tomorrow
        tomorrow = datetime.utcnow() + timedelta(days=1)
        tomorrow_start = tomorrow.replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow_end = tomorrow.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        forms = Form.query.filter(
            Form.admin_id == admin_id,
            Form.is_active == True,
            Form.deadline >= tomorrow_start,
            Form.deadline <= tomorrow_end
        ).all()
        
        if not forms:
            return jsonify({'message': 'No forms with deadlines tomorrow'}), 200
        
        sent_count = 0
        for form in forms:
            # Get students in the same branch who haven't responded
            students = Student.query.filter_by(branch=form.branch).all()
            
            for student in students:
                # Check if student has already responded
                existing_response = FormResponse.query.filter_by(
                    form_id=form.id,
                    student_id=student.id
                ).first()
                
                if not existing_response:
                    # Send reminder email
                    try:
                        msg = Message(
                            subject=f'Form Deadline Reminder: {form.title}',
                            sender=app.config['MAIL_DEFAULT_SENDER'],
                            recipients=[student.email]
                        )
                        
                        deadline_str = form.deadline.strftime('%B %d, %Y at %I:%M %p')
                        
                        msg.body = f"""
Dear {student.name},

This is a friendly reminder that you have a pending form submission with a deadline tomorrow.

Form Details:
- Title: {form.title}
- Description: {form.description}
- Deadline: {deadline_str}

Please log in to your student dashboard and complete this form before the deadline.

If you have any questions, please contact your department admin.

Best regards,
SAT Portal Team
                        """
                        
                        mail.send(msg)
                        sent_count += 1
                        
                    except Exception as e:
                        print(f"Failed to send reminder to {student.email}: {str(e)}")
                        continue
        
        return jsonify({
            'message': f'Deadline reminders sent successfully',
            'forms_processed': len(forms),
            'emails_sent': sent_count
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== Admin Forms: Unsubmitted Students ==========
@app.route('/api/admin/forms/<int:form_id>/unsubmitted', methods=['GET'])
def get_unsubmitted_students(form_id):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400

        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404

        form = db.session.get(Form, form_id)
        if not form:
            return jsonify({'error': 'Form not found'}), 404
        if form.admin_id != admin.id and ADMIN_CREDENTIALS.get('super@admin', {}).get('branch') != admin.branch:
            # Allow super admin to view all
            return jsonify({'error': 'Access denied'}), 403

        # Optional filters: year, section
        year = request.args.get('year')
        section = request.args.get('section')

        # Students in admin branch
        students_query = Student.query.filter_by(branch=admin.branch)
        if year:
            students_query = students_query.filter_by(year=year)
        if section:
            students_query = students_query.filter_by(section=section)
        all_students = students_query.all()

        # Responded student ids
        responded = FormResponse.query.with_entities(FormResponse.student_id).filter_by(form_id=form_id).all()
        responded_ids = {sid for (sid,) in responded}

        unsubmitted = [s for s in all_students if s.id not in responded_ids]

        data = []
        for s in unsubmitted:
            data.append({
                'id': s.id,
                'name': s.name,
                'rollnumber': s.rollnumber,
                'email': s.email,
                'phone': s.phone,
                'branch': s.branch,
                'section': s.section,
                'year': s.year,
                'gender': s.gender
            })

        return jsonify({'unsubmitted': data, 'count': len(data)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/<int:form_id>/unsubmitted/download', methods=['GET'])
def download_unsubmitted_students_excel(form_id):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400

        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404

        form = db.session.get(Form, form_id)
        if not form:
            return jsonify({'error': 'Form not found'}), 404
        if form.admin_id != admin.id and ADMIN_CREDENTIALS.get('super@admin', {}).get('branch') != admin.branch:
            return jsonify({'error': 'Access denied'}), 403

        # Filters
        year = request.args.get('year')
        section = request.args.get('section')

        students_query = Student.query.filter_by(branch=admin.branch)
        if year:
            students_query = students_query.filter_by(year=year)
        if section:
            students_query = students_query.filter_by(section=section)
        all_students = students_query.all()

        responded = FormResponse.query.with_entities(FormResponse.student_id).filter_by(form_id=form_id).all()
        responded_ids = {sid for (sid,) in responded}
        unsubmitted = [s for s in all_students if s.id not in responded_ids]

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unsubmitted Students"

        headers = ['ID', 'Name', 'Roll Number', 'Email', 'Phone', 'Branch', 'Section', 'Year', 'Gender']
        ws.append(headers)

        for s in unsubmitted:
            ws.append([
                s.id,
                s.name,
                s.rollnumber,
                s.email,
                s.phone or 'N/A',
                s.branch or 'N/A',
                s.section or 'N/A',
                s.year or 'N/A',
                s.gender or 'N/A'
            ])

        # Autosize columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'unsubmitted_students_form_{form_id}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== Super Admin: Manage Admins and Messages ==========
@app.route('/api/superadmin/admins', methods=['GET'])
def superadmin_list_admins():
    try:
        admins = Admin.query.order_by(Admin.branch.asc()).all()
        data = []
        for a in admins:
            data.append({
                'id': a.id,
                'name': a.name,
                'employee_id': a.employee_id,
                'email': a.email,
                'branch': a.branch
            })
        return jsonify({'admins': data}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/superadmin/admins/<int:admin_id>/password', methods=['PUT'])
def superadmin_change_admin_password(admin_id):
    try:
        data = request.get_json()
        new_password = data.get('new_password')
        if not new_password:
            return jsonify({'error': 'New password is required'}), 400
        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        # Update admin password in database
        admin.password = hash_password(new_password)
        
        # Update the password in ADMIN_CREDENTIALS for immediate effect
        if admin.employee_id in ADMIN_CREDENTIALS:
            ADMIN_CREDENTIALS[admin.employee_id]['password'] = new_password
        
        db.session.commit()
        return jsonify({'message': 'Admin password updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/superadmin/admins/<int:admin_id>', methods=['DELETE'])
def superadmin_delete_admin(admin_id):
    try:
        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        db.session.delete(admin)
        db.session.commit()
        return jsonify({'message': 'Admin deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/superadmin/messages', methods=['POST'])
def superadmin_send_message():
    try:
        data = request.get_json()
        admin_id = data.get('admin_id')
        subject = data.get('subject')
        body = data.get('body')
        if not all([admin_id, subject, body]):
            return jsonify({'error': 'admin_id, subject, and body are required'}), 400
        admin = db.session.get(Admin, admin_id)
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        msg = AdminMessage(admin_id=admin_id, subject=subject, body=body)
        db.session.add(msg)
        db.session.commit()
        return jsonify({'message': 'Message sent to admin'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/messages', methods=['GET'])
def admin_get_messages():
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        messages = AdminMessage.query.filter_by(admin_id=admin_id).order_by(AdminMessage.created_at.desc()).all()
        unread_count = AdminMessage.query.filter_by(admin_id=admin_id, is_read=False).count()
        data = []
        for m in messages:
            data.append({
                'id': m.id,
                'subject': m.subject,
                'body': m.body,
                'is_read': m.is_read,
                'created_at': m.created_at.isoformat()
            })
        return jsonify({'messages': data, 'unread_count': unread_count}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/messages/<int:message_id>/read', methods=['PUT'])
def admin_mark_message_read(message_id):
    try:
        msg = db.session.get(AdminMessage, message_id)
        if not msg:
            return jsonify({'error': 'Message not found'}), 404
        msg.is_read = True
        db.session.commit()
        return jsonify({'message': 'Message marked as read'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/forms/<int:form_id>', methods=['DELETE'])
def delete_form(form_id):
    try:
        admin_id = request.args.get('admin_id')
        if not admin_id:
            return jsonify({'error': 'Admin ID required'}), 400
        
        # Get the form and verify admin ownership
        form = Form.query.filter_by(id=form_id, admin_id=admin_id).first()
        if not form:
            return jsonify({'error': 'Form not found or access denied'}), 404
        
        # Get all responses for this form
        responses = FormResponse.query.filter_by(form_id=form_id).all()
        
        # Delete uploaded files for each response
        for response in responses:
            response_dir = os.path.join('uploads', 'forms', str(form_id), str(response.id))
            if os.path.exists(response_dir):
                try:
                    # Remove all files in the response directory
                    for filename in os.listdir(response_dir):
                        file_path = os.path.join(response_dir, filename)
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                    # Remove the response directory
                    os.rmdir(response_dir)
                except Exception as e:
                    print(f"Error deleting files for response {response.id}: {str(e)}")
        
        # Delete the form directory (including temp files)
        form_dir = os.path.join('uploads', 'forms', str(form_id))
        if os.path.exists(form_dir):
            try:
                # Remove all files and subdirectories
                for root, dirs, files in os.walk(form_dir, topdown=False):
                    for file in files:
                        os.remove(os.path.join(root, file))
                    for dir in dirs:
                        os.rmdir(os.path.join(root, dir))
                # Remove the form directory
                os.rmdir(form_dir)
            except Exception as e:
                print(f"Error deleting form directory: {str(e)}")
        
        # Delete all form responses from database
        FormResponse.query.filter_by(form_id=form_id).delete()
        
        # Delete the form from database
        db.session.delete(form)
        db.session.commit()
        
        return jsonify({'message': 'Form and all responses deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        sync_admin_credentials()
    app.run(debug=True, host='0.0.0.0', port=5000) 